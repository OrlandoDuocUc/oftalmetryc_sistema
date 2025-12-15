# adapters/input/flask_app/controllers/product_controller.py

from flask import Blueprint, render_template, redirect, url_for, request, session, flash
from app.domain.use_cases.product_use_cases import ProductUseCases
from app.domain.use_cases.services.sale_service import SaleService

product_html = Blueprint('product_html', __name__)

# Instancias de servicio para el controlador
product_use_cases = ProductUseCases()
sale_service = SaleService()

@product_html.route('/dashboard')  # mejor explícito que '/'
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('user_html.login'))

    # Valores por defecto
    stock_total_productos = 0
    productos_bajo_stock = []
    total_ventas_hoy = 0
    ventas_recientes = []

    try:
        # Si el servicio trae el método, úsalo
        if hasattr(sale_service, "get_dashboard_stats"):
            stats = sale_service.get_dashboard_stats() or {}
            stock_total_productos = stats.get('total_stock', 0)
            productos_bajo_stock = stats.get('productos_stock_bajo', [])
            total_ventas_hoy = stats.get('total_ventas_hoy', 0)
            ventas_recientes = stats.get('ventas_recientes', [])
        else:
            # Fallback mínimo sin romper la vista
            productos = product_use_cases.list_products()
            stock_total_productos = sum(getattr(p, "stock", 0) or 0 for p in productos)
            productos_bajo_stock = [p for p in productos if (getattr(p, "stock", 0) or 0) <= 10]
    except Exception as e:
        print(f"Error en dashboard: {e}")
        flash('Ocurrió un error al cargar el dashboard.', 'danger')

    return render_template(
        'dashboard.html',
        stock_total_productos=stock_total_productos,
        productos_bajo_stock=productos_bajo_stock,
        total_ventas_hoy=total_ventas_hoy,
        ventas_recientes=ventas_recientes,
        pagos_pendientes=0
    )

@product_html.route('/productos', methods=['GET', 'POST'])
def productos():
    if session.get('rol', '').lower() != 'administrador':
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('user_html.login'))

    if request.method == 'POST':
        try:
            from datetime import datetime
            data = request.form.to_dict()
            
            # Convertir tipos de datos
            if 'fecha' in data and data['fecha']:
                data['fecha'] = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
            else:
                data['fecha'] = datetime.now().date()
            
            # Campos requeridos
            data['cantidad'] = int(data.get('cantidad', 0))
            data['costo_unitario'] = float(data.get('costo_unitario', 0.0))
            
            # Costo total: calcular automáticamente si no viene o está vacío
            if data.get('costo_total') and data.get('costo_total').strip():
                data['costo_total'] = float(data['costo_total'])
            else:
                data['costo_total'] = data['cantidad'] * data['costo_unitario']
            
            # Campos numéricos opcionales - validar que no sean cadenas vacías
            if data.get('costo_venta_1') and data.get('costo_venta_1').strip():
                data['costo_venta_1'] = float(data['costo_venta_1'])
            else:
                data['costo_venta_1'] = None
                
            if data.get('costo_venta_2') and data.get('costo_venta_2').strip():
                data['costo_venta_2'] = float(data['costo_venta_2'])
            else:
                data['costo_venta_2'] = None
            
            if data.get('diametro_1') and data.get('diametro_1').strip():
                data['diametro_1'] = int(data['diametro_1'])
            else:
                data['diametro_1'] = None
            
            if data.get('diametro_2') and data.get('diametro_2').strip():
                data['diametro_2'] = int(data['diametro_2'])
            else:
                data['diametro_2'] = None

            product_use_cases.create_product(data)
            print(f"✅ Producto creado exitosamente: {data.get('nombre')}")
            flash('Producto creado exitosamente.', 'success')
            return redirect(url_for('product_html.productos', created='true'))
        except Exception as e:
            print(f"❌ Error al crear producto: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Ocurrió un error al crear el producto: {e}', 'danger')
            return redirect(url_for('product_html.productos'))

    # GET
    try:
        products = product_use_cases.list_products()
        return render_template('productos.html', products=products)
    except Exception as e:
        print(f"Error al listar productos: {e}")
        flash('Ocurrió un error al cargar los productos.', 'danger')
        return render_template('productos.html', products=[])

@product_html.route('/productos/eliminados')
def productos_eliminados():
    if session.get('rol', '').lower() != 'administrador':
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('user_html.login'))

    try:
        deleted_products = product_use_cases.list_deleted_products()
        return render_template('productos.html', products=[], deleted_products=deleted_products)
    except Exception as e:
        print(f"Error al listar productos eliminados: {e}")
        flash('Ocurrió un error al cargar los productos eliminados.', 'danger')
        return render_template('productos.html', products=[], deleted_products=[])

@product_html.route('/productos/editar/<int:product_id>', methods=['POST'])
def editar_producto(product_id):
    if session.get('rol', '').lower() != 'administrador':
        return redirect(url_for('user_html.login'))
    try:
        from datetime import datetime
        data = request.form.to_dict()
        
        # Convertir tipos de datos
        if 'fecha' in data and data['fecha']:
            data['fecha'] = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        
        # Campos requeridos
        data['cantidad'] = int(data.get('cantidad', 0))
        data['costo_unitario'] = float(data.get('costo_unitario', 0.0))
        
        # Costo total: calcular automáticamente si no viene o está vacío
        if data.get('costo_total') and data.get('costo_total').strip():
            data['costo_total'] = float(data['costo_total'])
        else:
            data['costo_total'] = data['cantidad'] * data['costo_unitario']
        
        # Campos numéricos opcionales - validar que no sean cadenas vacías
        if data.get('costo_venta_1') and data.get('costo_venta_1').strip():
            data['costo_venta_1'] = float(data['costo_venta_1'])
        else:
            data['costo_venta_1'] = None
            
        if data.get('costo_venta_2') and data.get('costo_venta_2').strip():
            data['costo_venta_2'] = float(data['costo_venta_2'])
        else:
            data['costo_venta_2'] = None
        
        if data.get('diametro_1') and data.get('diametro_1').strip():
            data['diametro_1'] = int(data['diametro_1'])
        else:
            data['diametro_1'] = None
        
        if data.get('diametro_2') and data.get('diametro_2').strip():
            data['diametro_2'] = int(data['diametro_2'])
        else:
            data['diametro_2'] = None

        product_use_cases.update_product(product_id, data)
        flash('Producto actualizado correctamente.', 'success')
    except Exception as e:
        print(f"Error al editar producto: {e}")
        flash(f'Ocurrió un error al editar el producto: {e}', 'danger')
    return redirect(url_for('product_html.productos'))

@product_html.route('/productos/eliminar/<int:product_id>', methods=['POST'])
def eliminar_producto(product_id):
    if session.get('rol', '').lower() != 'administrador':
        return redirect(url_for('user_html.login'))
    try:
        success = product_use_cases.delete_product(product_id)
        if success:
            flash('Producto eliminado físicamente.', 'success')
        else:
            flash('El producto no se eliminó físicamente (puede tener ventas asociadas) y fue desactivado.', 'warning')
    except Exception as e:
        print(f"Error al eliminar producto: {e}")
        flash('Ocurrió un error al eliminar el producto.', 'danger')
    return redirect(url_for('product_html.productos'))

@product_html.route('/productos/restore/<int:product_id>', methods=['POST'])
def restaurar_producto(product_id):
    if session.get('rol', '').lower() != 'administrador':
        return redirect(url_for('user_html.login'))
    try:
        product_use_cases.restore_product(product_id)
        flash('Producto restaurado exitosamente.', 'success')
    except Exception as e:
        print(f"Error al restaurar producto: {e}")
        flash('Ocurrió un error al restaurar el producto.', 'danger')
    return redirect(url_for('product_html.productos_eliminados'))

@product_html.route('/inventario/exportar-excel', methods=['POST'])
def exportar_inventario_excel():
    """Genera un archivo Excel con el inventario actual"""
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import send_file
        
        data = request.get_json()
        productos = data.get('productos', [])
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Encabezados con estilo
        headers = ['Fecha', 'Nombre', 'Distribuidor', 'Marca', 'Material', 'Tipo Armazón', 
                  'Código', 'Diámetro 1', 'Diámetro 2', 'Color', 'Cantidad', 
                  'Costo Unitario', 'Costo Total', 'Venta 1', 'Venta 2', 'Estado']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Agregar datos
        for row_num, producto in enumerate(productos, 2):
            ws.cell(row=row_num, column=1, value=producto['fecha'])
            ws.cell(row=row_num, column=2, value=producto['nombre'])
            ws.cell(row=row_num, column=3, value=producto['distribuidor'])
            ws.cell(row=row_num, column=4, value=producto['marca'])
            ws.cell(row=row_num, column=5, value=producto['material'])
            ws.cell(row=row_num, column=6, value=producto['tipo_armazon'])
            ws.cell(row=row_num, column=7, value=producto['codigo'])
            ws.cell(row=row_num, column=8, value=producto['diametro_1'])
            ws.cell(row=row_num, column=9, value=producto['diametro_2'])
            ws.cell(row=row_num, column=10, value=producto['color'])
            ws.cell(row=row_num, column=11, value=producto['cantidad'])
            ws.cell(row=row_num, column=12, value=producto['costo_unitario'])
            ws.cell(row=row_num, column=13, value=producto['costo_total'])
            ws.cell(row=row_num, column=14, value=producto['costo_venta_1'])
            ws.cell(row=row_num, column=15, value=producto['costo_venta_2'])
            ws.cell(row=row_num, column=16, value=producto['estado'])
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventario_{request.args.get("fecha", "")}.xlsx'
        )
    except Exception as e:
        print(f"Error al exportar inventario: {e}")
        return {'error': str(e)}, 500
