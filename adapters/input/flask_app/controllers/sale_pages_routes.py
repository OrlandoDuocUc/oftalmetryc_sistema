# adapters/input/flask_app/controllers/sale_pages_routes.py

from flask import Blueprint, render_template, request, redirect, url_for, session, flash
import json
from sqlalchemy.orm import joinedload

from app.infraestructure.utils.db import SessionLocal
from app.infraestructure.repositories.sql_product_repository import SQLProductRepository
from app.domain.use_cases.services.sale_service import SaleService
from app.domain.models.sale import Sale, SaleDetail

sale_html = Blueprint("sale_html", __name__)

sale_service = SaleService()

# GET: pantalla para armar el pedido (usa registrar_venta.html)
@sale_html.route("/registrar-venta", methods=["GET"])
def registrar_venta_page():
    with SessionLocal() as db:
        products = SQLProductRepository(db).get_all(include_deleted=False)
        products_json = json.dumps([
            {
                "producto_id": p.producto_id,
                "nombre": p.nombre,
                "precio_unitario": float(p.precio_unitario or 0)
            }
            for p in products
        ])
    return render_template("registrar_venta.html", products_json=products_json)


# POST: recibe los items del formulario anterior y muestra confirmar_venta.html
@sale_html.route("/revisar-venta", methods=["POST"])
def revisar_venta_page():
    raw = request.form.get("pedido_items", "")
    try:
        items = json.loads(raw) if raw else []
    except Exception:
        items = []

    if not items:
        flash("Debes agregar al menos un producto al pedido.", "warning")
        return redirect(url_for("sale_html.registrar_venta_page"))

    total_general = sum(float(i.get("subtotal", 0)) for i in items)

    # Guardamos en sesión para usarlos al finalizar
    session["venta_items"] = items
    session["venta_total"] = total_general

    return render_template("confirmar_venta.html", items=items, total_general=total_general)


# POST: guarda cliente (si viene) + registra venta + detalles + descuenta stock y redirige a boleta
@sale_html.route("/finalizar-venta-definitiva", methods=["POST"])
def finalizar_venta_definitiva():
    items = session.get("venta_items", [])
    total_general = session.get("venta_total", 0.0)

    if not items:
        flash("No hay items en la venta. Vuelve a armar el pedido.", "warning")
        return redirect(url_for("sale_html.registrar_venta_page"))

    # Datos de cliente desde el form (agregamos RUT porque en tu DB es NOT NULL si lo creas)
    cliente_data = {
        "rut": (request.form.get("cliente_rut") or "").strip(),
        "nombres": (request.form.get("cliente_nombres") or "").strip(),
        "ap_pat": (request.form.get("cliente_ap_pat") or "").strip(),
        "ap_mat": (request.form.get("cliente_ap_mat") or "").strip(),
        "telefono": (request.form.get("cliente_telefono") or "").strip(),
        "email": (request.form.get("cliente_email") or "").strip(),
        "direccion": (request.form.get("cliente_direccion") or "").strip(),
    }

    # Si no hay RUT, NO intentamos crear cliente (tu DB lo exige NOT NULL).
    if not cliente_data["rut"]:
        cliente_data = {}  # no se creará cliente

    # Usuario actual (ajústalo si guardas el id en la sesión)
    usuario_id = session.get("usuario_id", 1)

    try:
        venta_id = sale_service.register_sale_from_cart(
            cart_items=items,
            usuario_id=usuario_id,
            cliente_data=cliente_data,
            metodo_pago="efectivo",
            observaciones=None,
            descuento=0
        )
    except Exception as e:
        flash(f"Error al finalizar la venta definitiva: {e}", "danger")
        return redirect(url_for("sale_html.registrar_venta_page"))

    # limpiar sesión de la venta
    session.pop("venta_items", None)
    session.pop("venta_total", None)

    return redirect(url_for("sale_html.boleta_page", venta_id=venta_id))


# GET: boleta
@sale_html.route("/boleta/<int:venta_id>", methods=["GET"])
def boleta_page(venta_id: int):
    with SessionLocal() as db:
        venta = (
            db.query(Sale)
            .options(
                joinedload(Sale.cliente),
                joinedload(Sale.detalles).joinedload(SaleDetail.producto)
            )
            .filter(Sale.venta_id == venta_id)
            .first()
        )

        if not venta:
            flash("Venta no encontrada.", "warning")
            return redirect(url_for("sale_html.registrar_venta_page"))

    return render_template("boleta.html", venta=venta)


# (Opcional) Historial de ventas
@sale_html.route("/historial-ventas", methods=["GET"])
def historial_ventas_page():
    with SessionLocal() as db:
        ventas = (
            db.query(Sale)
            .options(joinedload(Sale.cliente))
            .order_by(Sale.fecha_venta.desc())
            .all()
        )
    return render_template("historial_ventas.html", ventas=ventas)

@sale_html.route("/ventas/exportar-excel", methods=["POST"])
def exportar_ventas_excel():
    """Genera un archivo Excel con el historial de ventas"""
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import send_file
        
        data = request.get_json()
        ventas = data.get('ventas', [])
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Ventas"
        
        # Encabezados con estilo
        headers = ['ID Venta', 'Producto', 'Cantidad', 'Total', 'Cliente', 'Vendedor', 'Fecha']
        
        header_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Agregar datos
        for row_num, venta in enumerate(ventas, 2):
            ws.cell(row=row_num, column=1, value=venta['id_venta'])
            ws.cell(row=row_num, column=2, value=venta['producto'])
            ws.cell(row=row_num, column=3, value=venta['cantidad'])
            ws.cell(row=row_num, column=4, value=venta['total'])
            ws.cell(row=row_num, column=5, value=venta['cliente'])
            ws.cell(row=row_num, column=6, value=venta['vendedor'])
            ws.cell(row=row_num, column=7, value=venta['fecha'])
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'ventas_{request.args.get("fecha", "")}.xlsx'
        )
    except Exception as e:
        print(f"Error al exportar ventas: {e}")
        return {'error': str(e)}, 500
